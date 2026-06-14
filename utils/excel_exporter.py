"""
Generates automated Excel reports with:
- Pivot Tables, Charts, Conditional Formatting
- KPI Summary sheet
- Data Quality sheet
"""
import io
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime


HEADER_FILL   = PatternFill("solid", fgColor="1A1A2E")
ACCENT_FILL   = PatternFill("solid", fgColor="16213E")
ALT_ROW_FILL  = PatternFill("solid", fgColor="F0F4FF")
KPI_FILL      = PatternFill("solid", fgColor="0F3460")
GREEN_FILL    = PatternFill("solid", fgColor="D4EDDA")
RED_FILL      = PatternFill("solid", fgColor="F8D7DA")

WHITE_BOLD    = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
DARK_BOLD     = Font(color="1A1A2E", bold=True, size=11, name="Calibri")
KPI_FONT      = Font(color="FFFFFF", bold=True, size=18, name="Calibri")
KPI_LBL_FONT  = Font(color="E0E0E0", size=9, name="Calibri")
CENTER        = Alignment(horizontal="center", vertical="center")
THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header_row(ws, row, cols, fill=HEADER_FILL, font=WHITE_BOLD):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _auto_width(ws, min_width=12):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len, 40)


def generate_excel_report(df: pd.DataFrame, kpis: dict, issues: list) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: KPI Summary ─────────────────────────────────
    ws_kpi = wb.active
    ws_kpi.title = "KPI Summary"
    ws_kpi.sheet_view.showGridLines = False

    # Title banner
    ws_kpi.merge_cells("A1:F1")
    ws_kpi["A1"] = "📊 Business Reporting Dashboard — KPI Summary"
    ws_kpi["A1"].fill = PatternFill("solid", fgColor="0F3460")
    ws_kpi["A1"].font = Font(color="FFFFFF", bold=True, size=16, name="Calibri")
    ws_kpi["A1"].alignment = CENTER
    ws_kpi.row_dimensions[1].height = 40

    ws_kpi.merge_cells("A2:F2")
    ws_kpi["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    ws_kpi["A2"].font = Font(color="888888", italic=True, size=9)
    ws_kpi["A2"].alignment = CENTER

    # KPI cards (row 4–7)
    kpi_items = [
        ("💰 Total Revenue", f"₹{kpis['total_revenue']:,.0f}", "A"),
        ("📦 Total Orders", f"{kpis['total_orders']:,}", "B"),
        ("🛒 Avg Order Value", f"₹{kpis['avg_order_value']:,.0f}", "C"),
        ("❌ Cancellation Rate", f"{kpis['cancellation_rate']}%", "D"),
        ("👥 Unique Customers", f"{kpis['unique_customers']:,}", "E"),
        ("✅ Completed Orders", f"{kpis['completed_orders']:,}", "F"),
    ]

    for label, value, col in kpi_items:
        ws_kpi.merge_cells(f"{col}4:{col}5")
        ws_kpi.merge_cells(f"{col}6:{col}7")
        lbl_cell = ws_kpi[f"{col}4"]
        val_cell = ws_kpi[f"{col}6"]
        lbl_cell.value = label
        lbl_cell.fill = KPI_FILL
        lbl_cell.font = KPI_LBL_FONT
        lbl_cell.alignment = CENTER
        val_cell.value = value
        val_cell.fill = KPI_FILL
        val_cell.font = KPI_FONT
        val_cell.alignment = CENTER

    ws_kpi.row_dimensions[4].height = 20
    ws_kpi.row_dimensions[5].height = 20
    ws_kpi.row_dimensions[6].height = 30
    ws_kpi.row_dimensions[7].height = 30

    # Category table
    cat_df = df[df["status"] == "Completed"].groupby("category").agg(
        Total_Orders=("order_id", "count"),
        Total_Revenue=("revenue", "sum"),
        Avg_Order_Value=("revenue", "mean"),
        Units_Sold=("quantity", "sum"),
    ).reset_index().sort_values("Total_Revenue", ascending=False)

    ws_kpi["A9"] = "Category Performance"
    ws_kpi["A9"].font = Font(bold=True, size=12, color="0F3460")
    ws_kpi.row_dimensions[9].height = 20

    headers = ["Category", "Total Orders", "Total Revenue (₹)", "Avg Order Value (₹)", "Units Sold"]
    for c, h in enumerate(headers, 1):
        ws_kpi.cell(row=10, column=c, value=h)
    _style_header_row(ws_kpi, 10, len(headers))

    for r, row in enumerate(cat_df.itertuples(index=False), 11):
        ws_kpi.cell(row=r, column=1, value=row.category)
        ws_kpi.cell(row=r, column=2, value=row.Total_Orders)
        ws_kpi.cell(row=r, column=3, value=round(row.Total_Revenue, 2))
        ws_kpi.cell(row=r, column=4, value=round(row.Avg_Order_Value, 2))
        ws_kpi.cell(row=r, column=5, value=row.Units_Sold)
        if r % 2 == 0:
            for c in range(1, 6):
                ws_kpi.cell(row=r, column=c).fill = ALT_ROW_FILL

    # Color scale on revenue column
    last_data_row = 10 + len(cat_df)
    ws_kpi.conditional_formatting.add(
        f"C11:C{last_data_row}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="0F3460")
    )

    _auto_width(ws_kpi)

    # ── Sheet 2: Revenue Trend ────────────────────────────────
    ws_trend = wb.create_sheet("Revenue Trend")
    ws_trend.sheet_view.showGridLines = False

    trend_df = df[df["status"] == "Completed"].copy()
    trend_df["Month"] = trend_df["order_date"].dt.to_period("M").astype(str)
    monthly = trend_df.groupby("Month").agg(
        Revenue=("revenue", "sum"),
        Orders=("order_id", "count")
    ).reset_index()

    ws_trend["A1"] = "Monthly Revenue Trend"
    ws_trend["A1"].font = Font(bold=True, size=14, color="0F3460")

    for c, h in enumerate(["Month", "Total Revenue (₹)", "Total Orders"], 1):
        ws_trend.cell(row=2, column=c, value=h)
    _style_header_row(ws_trend, 2, 3)

    for r, row in enumerate(monthly.itertuples(index=False), 3):
        ws_trend.cell(row=r, column=1, value=row.Month)
        ws_trend.cell(row=r, column=2, value=round(row.Revenue, 2))
        ws_trend.cell(row=r, column=3, value=row.Orders)

    # Line chart
    chart = LineChart()
    chart.title = "Monthly Revenue"
    chart.style = 10
    chart.y_axis.title = "Revenue (₹)"
    chart.x_axis.title = "Month"
    chart.width = 22
    chart.height = 12
    data = Reference(ws_trend, min_col=2, min_row=2, max_row=2 + len(monthly))
    cats = Reference(ws_trend, min_col=1, min_row=3, max_row=2 + len(monthly))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_trend.add_chart(chart, "E2")
    _auto_width(ws_trend)

    # ── Sheet 3: Regional Analysis ────────────────────────────
    ws_region = wb.create_sheet("Regional Analysis")
    ws_region.sheet_view.showGridLines = False

    region_df = df.groupby("region").agg(
        Total_Orders=("order_id", "count"),
        Total_Revenue=("revenue", "sum"),
        Unique_Customers=("customer_id", "nunique"),
    ).reset_index().sort_values("Total_Revenue", ascending=False)

    ws_region["A1"] = "Regional Sales Analysis"
    ws_region["A1"].font = Font(bold=True, size=14, color="0F3460")

    for c, h in enumerate(["Region", "Total Orders", "Total Revenue (₹)", "Unique Customers"], 1):
        ws_region.cell(row=2, column=c, value=h)
    _style_header_row(ws_region, 2, 4)

    for r, row in enumerate(region_df.itertuples(index=False), 3):
        ws_region.cell(row=r, column=1, value=row.region)
        ws_region.cell(row=r, column=2, value=row.Total_Orders)
        ws_region.cell(row=r, column=3, value=round(row.Total_Revenue, 2))
        ws_region.cell(row=r, column=4, value=row.Unique_Customers)

    # Bar chart
    bar = BarChart()
    bar.type = "col"
    bar.title = "Revenue by Region"
    bar.style = 10
    bar.width = 18
    bar.height = 12
    data = Reference(ws_region, min_col=3, min_row=2, max_row=2 + len(region_df))
    cats = Reference(ws_region, min_col=1, min_row=3, max_row=2 + len(region_df))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws_region.add_chart(bar, "F2")
    _auto_width(ws_region)

    # ── Sheet 4: Data Quality ─────────────────────────────────
    ws_dq = wb.create_sheet("Data Quality")
    ws_dq.sheet_view.showGridLines = False

    ws_dq["A1"] = "Data Quality Report"
    ws_dq["A1"].font = Font(bold=True, size=14, color="0F3460")
    ws_dq["A2"] = f"Last Refreshed: {datetime.now().strftime('%d %b %Y %H:%M:%S')}"
    ws_dq["A2"].font = Font(italic=True, color="888888", size=9)

    for c, h in enumerate(["Issue Type", "Count", "Action Taken"], 1):
        ws_dq.cell(row=4, column=c, value=h)
    _style_header_row(ws_dq, 4, 3)

    for r, issue in enumerate(issues, 5):
        ws_dq.cell(row=r, column=1, value=issue["type"])
        ws_dq.cell(row=r, column=2, value=issue["count"])
        ws_dq.cell(row=r, column=3, value=issue["action"])
        fill = RED_FILL if issue["count"] > 0 and "Records" not in issue["type"] else GREEN_FILL
        for c in range(1, 4):
            ws_dq.cell(row=r, column=c).fill = fill

    _auto_width(ws_dq)

    # ── Sheet 5: Raw Data ─────────────────────────────────────
    ws_raw = wb.create_sheet("Raw Data")
    ws_raw.sheet_view.showGridLines = False
    ws_raw["A1"] = "Cleaned Orders Dataset"
    ws_raw["A1"].font = Font(bold=True, size=14, color="0F3460")

    raw_cols = ["order_id", "order_date", "category", "region",
                "payment_method", "status", "revenue", "quantity", "customer_id"]
    for c, h in enumerate(raw_cols, 1):
        ws_raw.cell(row=2, column=c, value=h.replace("_", " ").title())
    _style_header_row(ws_raw, 2, len(raw_cols))

    for r, row in enumerate(df[raw_cols].itertuples(index=False), 3):
        for c, val in enumerate(row, 1):
            cell = ws_raw.cell(row=r, column=c, value=val)
            if r % 2 == 0:
                cell.fill = ALT_ROW_FILL

    _auto_width(ws_raw)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
